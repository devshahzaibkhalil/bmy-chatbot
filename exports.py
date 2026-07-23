"""
exports.py
CSV / JSON / Excel / PDF export helpers for the admin dashboard.
All generated locally - openpyxl for Excel, reportlab for PDF, stdlib csv/json
for the rest. No external service involved.
"""

import csv
import io
import json

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

CONVERSATION_COLUMNS = [
    "id", "full_name", "email", "phone", "company_name",
    "interested_service", "budget", "timeline", "status",
    "started_at", "ended_at", "duration_seconds",
]

LEAD_COLUMNS = [
    "id", "name", "email", "phone", "company_name",
    "interested_service", "budget", "timeline", "status", "created_at",
]


def _rows_for(records, columns):
    return [{c: r.get(c, "") for c in columns} for r in records]


def to_csv_bytes(records, columns):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns)
    writer.writeheader()
    for row in _rows_for(records, columns):
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def to_json_bytes(records):
    return json.dumps(records, indent=2, default=str).encode("utf-8")


def to_excel_bytes(records, columns, sheet_title="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(columns)
    for row in _rows_for(records, columns):
        ws.append([row.get(c, "") for c in columns])

    # Light formatting: bold header, autosize-ish column widths
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for i, col in enumerate(columns, start=1):
        max_len = max([len(str(col))] + [len(str(r.get(col, ""))) for r in _rows_for(records, columns)])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf_bytes(records, columns, title="BMY Marketer - Export"):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    data = [columns] + [[str(r.get(c, "")) for c in columns] for r in _rows_for(records, columns)]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a4fd6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f3f9")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def export_conversations(records, fmt):
    fmt = fmt.lower()
    if fmt == "csv":
        return to_csv_bytes(records, CONVERSATION_COLUMNS), "text/csv", "conversations.csv"
    if fmt == "json":
        return to_json_bytes(records), "application/json", "conversations.json"
    if fmt == "xlsx" or fmt == "excel":
        return to_excel_bytes(records, CONVERSATION_COLUMNS, "Conversations"), \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "conversations.xlsx"
    if fmt == "pdf":
        return to_pdf_bytes(records, CONVERSATION_COLUMNS, "BMY Marketer - Conversations"), \
            "application/pdf", "conversations.pdf"
    raise ValueError(f"Unsupported export format: {fmt}")


def export_leads(records, fmt):
    fmt = fmt.lower()
    if fmt == "csv":
        return to_csv_bytes(records, LEAD_COLUMNS), "text/csv", "leads.csv"
    if fmt == "json":
        return to_json_bytes(records), "application/json", "leads.json"
    if fmt == "xlsx" or fmt == "excel":
        return to_excel_bytes(records, LEAD_COLUMNS, "Leads"), \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "leads.xlsx"
    if fmt == "pdf":
        return to_pdf_bytes(records, LEAD_COLUMNS, "BMY Marketer - Leads"), \
            "application/pdf", "leads.pdf"
    raise ValueError(f"Unsupported export format: {fmt}")
